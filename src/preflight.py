"""Evidence-only input discovery, field resolution and rule dependency preflight."""
from __future__ import annotations
import json
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
import pandas as pd
from src.field_mapping import FIELD_MAP

PROJECT_ROOT = Path(__file__).resolve().parents[1]
CONFIG = PROJECT_ROOT / "config"

def resolve_config(name: str, config_dir: Path | None = None) -> Path:
    """Resolve project configuration independently of the process cwd."""
    path = (config_dir if config_dir is not None else CONFIG) / name
    path = path.resolve()
    if not path.is_file():
        raise FileNotFoundError(f"config missing: resolved path={path}")
    return path

def _load(path: Path) -> dict[str, Any]:
    # Config files are JSON-compatible YAML, keeping the runtime dependency-free.
    return json.loads(path.read_text(encoding="utf-8"))

def _files(inputs: Path | list[Path]) -> list[Path]:
    paths = [inputs] if isinstance(inputs, Path) else inputs
    out = []
    for p in paths:
        if p.is_dir(): out.extend(sorted(x for x in p.glob("*.xlsx") if not x.name.startswith("~$"))); out.extend(sorted(p.glob("*.csv")))
        elif p.suffix.lower() in (".xlsx", ".csv"): out.append(p)
    return out

def _frame(path: Path, sheet: str | int | None, worksheet=None) -> tuple[str, pd.DataFrame]:
    if path.suffix.lower() == ".csv": return "__csv__", pd.read_csv(path, nrows=None)
    if worksheet is not None:
        rows = worksheet.iter_rows(values_only=True)
        header = next((row for row in rows if any(value is not None for value in row)), None)
        if header is None:
            return worksheet.title, pd.DataFrame()
        headers = [str(value) if value is not None else f"Unnamed: {index}" for index, value in enumerate(header)]
        return worksheet.title, pd.DataFrame(list(rows), columns=headers)
    chosen = sheet if sheet is not None else 0
    # Some real-world XLSX files omit <dimension>. In read_only mode openpyxl
    # then exposes max_row/max_column as None; iterate cells instead of using
    # those metadata fields to decide whether the worksheet is valid.
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[chosen] if isinstance(chosen, str) else workbook.worksheets[chosen]
    rows = worksheet.iter_rows(values_only=True)
    header = next((row for row in rows if any(value is not None for value in row)), None)
    if header is None:
        workbook.close()
        return worksheet.title, pd.DataFrame()
    headers = [str(value) if value is not None else f"Unnamed: {index}" for index, value in enumerate(header)]
    records = list(rows)
    frame = pd.DataFrame(records, columns=headers)
    workbook.close()
    return worksheet.title, frame

def resolve_fields(headers: list[str], source_file: str, source_sheet: str, config_dir: Path | None = None) -> list[dict[str, Any]]:
    aliases = _load(resolve_config("field_aliases.yaml", config_dir))["aliases"]
    reverse: dict[str, list[str]] = {}
    for canonical, names in aliases.items():
        for name in set(names + [FIELD_MAP[canonical].source]): reverse.setdefault(name, []).append(canonical)
    rows = []
    for raw in headers:
        candidates = reverse.get(str(raw).strip(), [])
        if len(candidates) == 1:
            canonical, method, status, confidence = candidates[0], ("exact" if raw == FIELD_MAP[candidates[0]].source else "alias"), "mapped", 1.0 if raw == FIELD_MAP[candidates[0]].source else .9
        elif len(candidates) > 1:
            canonical, method, status, confidence = None, "ambiguous", "ambiguous", None
        else:
            canonical, method, status, confidence = None, "unmapped", "unused", None
        rows.append({"source_file": source_file, "source_sheet": source_sheet, "raw_field": raw, "canonical_field": canonical, "mapping_method": method, "confidence": confidence, "used_by_metrics": [], "used_by_rules": [], "status": status})
    present = {r["canonical_field"] for r in rows if r["canonical_field"]}
    for canonical, spec in FIELD_MAP.items():
        if canonical not in present: rows.append({"source_file": source_file, "source_sheet": source_sheet, "raw_field": None, "canonical_field": canonical, "mapping_method": "none", "confidence": None, "used_by_metrics": [], "used_by_rules": [], "status": "missing"})
    return rows

def build_preflight(inputs: Path | list[Path], sheet: str | int | None = None, config_dir: Path | None = None) -> dict[str, Any]:
    metrics = _load(resolve_config("metric_dependencies.yaml", config_dir))["metrics"]
    rules = _load(resolve_config("rule_dependencies.yaml", config_dir))["rules"]
    files, discoveries, resolutions = _files(inputs), [], []
    complete_month_count = 0
    for path in files:
        workbook = None
        if path.suffix.lower() == ".csv":
            sheets = [sheet]
        else:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheets = list(workbook.sheetnames) if sheet is None else [sheet]
        for chosen in sheets:
            worksheet = None if workbook is None else (workbook[chosen] if isinstance(chosen, str) else workbook.worksheets[chosen])
            sheet_name, df = _frame(path, chosen, worksheet)
            dates = pd.to_datetime(df.get(FIELD_MAP["stat_date"].source), errors="coerce") if FIELD_MAP["stat_date"].source in df else pd.Series(dtype="datetime64[ns]")
            discoveries.append({"source_file": str(path), "source_sheet": sheet_name, "row_count": len(df), "headers": [str(x) for x in df.columns], "data_types": {str(k): str(v) for k,v in df.dtypes.items()}, "date_fields": [FIELD_MAP["stat_date"].source] if not dates.empty else [], "date_range": {"start": dates.min().date().isoformat(), "end": dates.max().date().isoformat()} if dates.notna().any() else None, "business_theme": "经营分析渠道/品类"})
            if dates.notna().any():
                by_month = dates.dropna().groupby(dates.dropna().dt.to_period("M"))
                complete_month_count = max(complete_month_count, sum(len(g) > 0 and g.min().day == 1 and g.max().day == g.max().days_in_month for _, g in by_month))
            resolutions.extend(resolve_fields([str(x) for x in df.columns], str(path), sheet_name, config_dir))
        if workbook is not None:
            workbook.close()
    for r in resolutions:
        c = r["canonical_field"]
        if c:
            r["used_by_metrics"] = [m for m,v in metrics.items() if c in v["fields"]]
            r["used_by_rules"] = [rid for rid,v in rules.items() if c in {field for metric in v["metrics"] for field in _metric_fields(metric)}]
    coverage=[]; gaps=[]
    complete_months = complete_month_count
    for rid, dep in rules.items():
        missing = [f for m in dep["metrics"] for f in _metric_fields(m) if not any(r["canonical_field"] == f and r["status"] == "mapped" for r in resolutions)]
        reasons = (["missing_fields:" + ",".join(missing)] if missing else []) + ([f"complete_months<{dep['min_complete_months']}"] if complete_months < dep["min_complete_months"] else [])
        coverage.append({"rule_id": rid, "rule_name": dep["name"], "status": "blocked" if reasons else "executable", "missing_dependencies": reasons, "screen": dep["screen"]})
        gaps.extend({"type": "rule_dependency", "rule_id": rid, "reason": x} for x in reasons)
    return {"schema_version":"1.0", "conclusion_generated":False, "discoveries":discoveries, "field_resolution":resolutions, "rule_coverage":coverage, "data_gaps":gaps}

def _metric_fields(name: str) -> list[str]:
    return {"category_revenue":["category","first_order_revenue"], "cac_rate":["cost","first_order_net_revenue"], "channel_roi":["cost","first_order_revenue"], "cost":["cost"]}.get(name, [name])

def write_preflight(payload: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True); path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="Evidence-only preflight; does not evaluate rules")
    parser.add_argument("input", type=Path)
    parser.add_argument("--sheet-name")
    parser.add_argument("--output", type=Path, default=Path("outputs/runtime/preflight.json"))
    args = parser.parse_args()
    write_preflight(build_preflight(args.input, args.sheet_name), args.output)
    print(f"preflight complete: {args.output}")

if __name__ == "__main__":
    main()
